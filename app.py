from fastapi import FastAPI, Request, Form, UploadFile, File
from fastapi.responses import HTMLResponse, RedirectResponse, JSONResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from pathlib import Path
from datetime import datetime
try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None
import sqlite3, shutil, csv, re

from pathlib import Path
import shutil

BASE_DIR = Path(__file__).resolve().parent

DB = BASE_DIR / "database.sqlite"

CARICAMENTI = BASE_DIR / "caricamenti"
CARICAMENTI.mkdir(exist_ok=True)

# I file sono nella cartella principale del repository
BONACCINI_CSV = BASE_DIR / "catalogo_bonaccini_v19.csv"
ORTO_CSV = BASE_DIR / "catalogo_ortofrutticola_categorie_v19.csv"
CONAD_XLSX = BASE_DIR / "PREZZI CONAD.xlsx"
GS_XLSX = BASE_DIR / "PREZZI GS.xlsx"

# Creiamo automaticamente la cartella static
STATIC_DIR = BASE_DIR / "static"
STATIC_DIR.mkdir(exist_ok=True)

# Copiamo CSS e JavaScript dalla cartella principale
for filename in ["style.css", "app.js"]:
    source = BASE_DIR / filename
    destination = STATIC_DIR / filename
    if source.exists():
        shutil.copy2(source, destination)

app = FastAPI(title="Gestione prezzi bar")

app.mount(
    "/static",
    StaticFiles(directory=str(STATIC_DIR)),
    name="static"
)

templates = Jinja2Templates(directory=str(BASE_DIR))


def conn():
    c = sqlite3.connect(DB)
    c.row_factory = sqlite3.Row
    return c


def to_float(v):
    if v is None:
        return None
    s = str(v).strip().replace("€", "").replace(" ", "").replace(",", ".")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def read_csv_rows(path):
    if not path.exists():
        return []
    with path.open(newline="", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))



def clean_text(s):
    return re.sub(r"\s+", " ", str(s or "").strip())

def infer_supplier_category(desc):
    s=clean_text(desc).upper()
    if "GIN" in s: return "Gin"
    if "AMARO" in s or "JAGERMEISTER" in s or "SAMBUCA" in s: return "Amari"
    if "APEROL" in s or "CAMPARI" in s or "BITTER" in s: return "Aperitivi/Bitter"
    if "VODKA" in s: return "Vodka"
    if "HAVANA" in s or "BACARDI" in s or "PAMPERO" in s or "RUM" in s: return "Rum"
    if "ACQUA" in s:
        v=volume_to_cl(None, s)
        if v is not None and abs(v-50)<0.01: return "Acqua 0,5 L"
        return "Acqua"
    if "RED BULL" in s: return "Energy drink"
    if "LEMONSODA" in s or "ORANSODA" in s: return "Bibite"
    if "SCHWEPPES" in s: return "Tonica"
    if "SUCCO" in s: return "Succhi"
    if "TEQUILA" in s: return "Tequila"
    if "PROSECCO" in s or "SPUMANTE" in s or "RIBOLLA" in s: return "Prosecco"
    if "VERMOUT" in s: return "Vermouth"
    if "GIUSTO SPIRITO" in s or "BIRRA" in s or "PILS" in s or "IPA" in s or "ROSSA" in s: return "Birre in fusto"
    return "Da classificare"

def volume_to_cl(v, desc=""):
    if v is not None:
        try: return float(v)
        except: pass
    s=clean_text(desc).upper().replace(",", ".")
    m=re.search(r"(\d+(?:\.\d+)?)\s*(CL|ML|LT|L)\b", s)
    if not m: return None
    n=float(m.group(1)); u=m.group(2)
    return n if u=="CL" else (n/10 if u=="ML" else n*100)

def import_core_supplier_products(cur):
    # Ortofrutticola is also mirrored into the generic supplier table so that
    # direct comparisons and same-category alternatives include it.
    for r in read_csv_rows(ORTO_CSV):
        desc=clean_text(r.get("descrizione"))
        if not desc:
            continue
        price=to_float(r.get("prezzo"))
        cat=clean_text(r.get("categoria")) or infer_supplier_category(desc)
        vol=volume_to_cl(None, desc)
        unit=(f"{int(vol)} cl" if vol is not None and float(vol).is_integer() else (f"{vol:g} cl" if vol is not None else clean_text(r.get("unita"))))
        existing=cur.execute(
            "SELECT id FROM supplier_products WHERE supplier=? AND descrizione=? AND ((volume_cl IS NULL AND ? IS NULL) OR volume_cl=?)",
            ("Ortofrutticola", desc, vol, vol)
        ).fetchone()
        if existing:
            cur.execute("""UPDATE supplier_products
                           SET categoria=?, prezzo=?, unita=?, source_file=?
                           WHERE id=?""",
                        (cat, price, unit, "catalogo_ortofrutticola_categorie_v19.csv", existing["id"]))
        else:
            cur.execute("""INSERT INTO supplier_products
                           (supplier,descrizione,categoria,prezzo,volume_cl,unita,source_file)
                           VALUES (?,?,?,?,?,?,?)""",
                        ("Ortofrutticola", desc, cat, price, vol, unit,
                         "catalogo_ortofrutticola_categorie_v19.csv"))

def import_supplier_excels(cur):
    if load_workbook is None: return
    if CONAD_XLSX.exists():
        wb=load_workbook(CONAD_XLSX,data_only=True,read_only=True)
        if "CONAD VOLUME SEPARATO" in wb.sheetnames:
            ws=wb["CONAD VOLUME SEPARATO"]
            for row in ws.iter_rows(values_only=True):
                vals=list(row)
                if len(vals)<4: continue
                desc=clean_text(vals[1]); vol=volume_to_cl(vals[2],desc); price=to_float(vals[3])
                if not desc or desc.upper()=="VOCE" or price is None: continue
                cat=infer_supplier_category(desc)
                cur.execute("""INSERT OR IGNORE INTO supplier_products
                (supplier,descrizione,categoria,prezzo,volume_cl,unita,source_file)
                VALUES (?,?,?,?,?,?,?)""",("Conad",desc,cat,price,vol,
                f"{int(vol)} cl" if vol and float(vol).is_integer() else (f"{vol} cl" if vol else ""),
                "PREZZI CONAD.xlsx / CONAD VOLUME SEPARATO"))
        wb.close()
    if GS_XLSX.exists():
        wb=load_workbook(GS_XLSX,data_only=True,read_only=True)
        ws=wb[wb.sheetnames[0]]
        for row in ws.iter_rows(values_only=True):
            vals=list(row)
            if len(vals)<4: continue
            desc=clean_text(vals[1]); liters=vals[2]; price=to_float(vals[3])
            if not desc or "GIUSTO SPIRITO" not in desc.upper() or price is None: continue
            vol=volume_to_cl(liters,desc)
            cur.execute("""INSERT OR IGNORE INTO supplier_products
            (supplier,descrizione,categoria,prezzo,volume_cl,unita,source_file)
            VALUES (?,?,?,?,?,?,?)""",("Giusto Spirito",desc,"Birre in fusto",price,vol,
            f"{liters} L" if liters else "","PREZZI GS.xlsx"))
        wb.close()

def init_db():
    CARICAMENTI.mkdir(exist_ok=True)
    BASE_DIR.mkdir(exist_ok=True)
    db = conn()
    cur = db.cursor()
    cur.execute("""CREATE TABLE IF NOT EXISTS bonaccini_products(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codice TEXT,
        descrizione TEXT NOT NULL,
        categoria TEXT NOT NULL DEFAULT 'Da classificare',
        prezzo REAL,
        unita TEXT,
        data_ultimo TEXT,
        quantita REAL DEFAULT 0,
        n_acquisti INTEGER DEFAULT 0,
        stato TEXT DEFAULT 'Solo Bonaccini'
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS ortofrutticola_products(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        codice TEXT,
        descrizione TEXT NOT NULL,
        categoria TEXT NOT NULL DEFAULT 'Da classificare',
        prezzo REAL,
        unita TEXT,
        pz_x_ct REAL
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS supplier_products(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        supplier TEXT NOT NULL,
        descrizione TEXT NOT NULL,
        categoria TEXT NOT NULL DEFAULT 'Da classificare',
        prezzo REAL,
        volume_cl REAL,
        unita TEXT DEFAULT '',
        source_file TEXT DEFAULT '',
        UNIQUE(supplier, descrizione, volume_cl)
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS direct_matches(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        bonaccini_id INTEGER NOT NULL,
        orto_id INTEGER NOT NULL,
        UNIQUE(bonaccini_id, orto_id)
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS orders(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        created_at TEXT NOT NULL,
        note TEXT
    )""")
    cur.execute("""CREATE TABLE IF NOT EXISTS order_lines(
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        order_id INTEGER,
        descrizione TEXT,
        categoria TEXT,
        quantita REAL,
        fornitore TEXT,
        prezzo REAL
    )""")

    b_rows = read_csv_rows(BONACCINI_CSV)
    o_rows = read_csv_rows(ORTO_CSV)

    # Se i CSV sono presenti, diventano la base operativa. Se cambia il numero righe, ricarica tutto.
    current_b = cur.execute("SELECT COUNT(*) FROM bonaccini_products").fetchone()[0]
    if b_rows and current_b != len(b_rows):
        cur.execute("DELETE FROM direct_matches")
        cur.execute("DELETE FROM bonaccini_products")
        for r in b_rows:
            cur.execute("""INSERT INTO bonaccini_products
                (codice, descrizione, categoria, prezzo, unita, data_ultimo, quantita, n_acquisti, stato)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)""", (
                r.get("codiceBonaccini") or r.get("codice"),
                (r.get("descrizione") or "").strip(),
                (r.get("categoria") or "Da classificare").strip() or "Da classificare",
                to_float(r.get("prezzoBonaccini") or r.get("prezzo")),
                r.get("umBonaccini") or r.get("unita") or "",
                r.get("dataUltimo") or "",
                to_float(r.get("quantitaAcquistata")) or 0,
                int(to_float(r.get("nAcquisti")) or 0),
                r.get("stato") or "Solo Bonaccini"
            ))

    current_o = cur.execute("SELECT COUNT(*) FROM ortofrutticola_products").fetchone()[0]
    if o_rows and current_o != len(o_rows):
        cur.execute("DELETE FROM direct_matches")
        cur.execute("DELETE FROM ortofrutticola_products")
        for r in o_rows:
            cur.execute("""INSERT INTO ortofrutticola_products
                (codice, descrizione, categoria, prezzo, unita, pz_x_ct)
                VALUES (?, ?, ?, ?, ?, ?)""", (
                r.get("codice"),
                (r.get("descrizione") or "").strip(),
                (r.get("categoria") or "Da classificare").strip() or "Da classificare",
                to_float(r.get("prezzo")),
                r.get("unita") or "",
                to_float(r.get("pz_x_ct"))
            ))

    import_core_supplier_products(cur)
    import_supplier_excels(cur)
    db.commit(); db.close()


@app.on_event("startup")
def startup():
    init_db()


def norm(s):
    return "".join(ch for ch in str(s).lower() if ch.isalnum())


def important_tokens(s):
    stop = {"gin","amaro","rum","vodka","birra","fusto","acqua","naturale","frizzante","cl","lt","ml","pet","lattina","bitter","premium","verum"}
    cleaned = str(s).replace(".", " ").replace("-", " ").replace(",", " ").upper()
    return [t for t in cleaned.split() if len(t) >= 5 and t.lower() not in stop and not any(ch.isdigit() for ch in t)]



def product_volume_cl(b):
    return volume_to_cl(b.get("unita"), b.get("descrizione"))

def supplier_matches(b):
    db=conn()
    rows=db.execute("SELECT * FROM supplier_products WHERE categoria=?", (b["categoria"],)).fetchall()
    db.close()
    tokens=important_tokens(b["descrizione"])
    bv=product_volume_cl(dict(b))
    out=[]
    for r0 in rows:
        r=dict(r0)
        score=sum(1 for t in tokens if t in r["descrizione"].upper())
        rv=r.get("volume_cl")
        if bv is not None and rv is not None:
            if abs(float(bv)-float(rv))<0.01: score+=2
            else: continue
        if score>0:
            r["_score"]=score; out.append(r)
    out.sort(key=lambda x:(-x["_score"], x["prezzo"] if x["prezzo"] is not None else 1e99))
    return out

def direct_matches_all(b):
    best={}
    for r in supplier_matches(b):
        if r["supplier"] not in best or r["_score"]>best[r["supplier"]]["_score"]:
            best[r["supplier"]]=r
    return best

def category_key(category):
    s=clean_text(category).lower().replace("à","a").replace("è","e")
    s=s.replace("0,5","0.5").replace("litri","l").replace("litro","l")
    s=re.sub(r"\\s+"," ",s)
    if "acqua" in s and ("0.5" in s or "50 cl" in s or "50cl" in s):
        return "acqua_05"
    if "birr" in s and ("fusto" in s or "21 l" in s or "20 l" in s or "30 l" in s):
        return "birra_fusto"
    if "gin" in s: return "gin"
    if "amar" in s: return "amari"
    if "rum" in s: return "rum"
    if "vodka" in s: return "vodka"
    if "prosecco" in s or "spumante" in s or "ribolla" in s: return "prosecco"
    return s

def alternatives_all(category, volume_cl=None):
    db=conn()
    rows=[dict(r) for r in db.execute("SELECT * FROM supplier_products ORDER BY supplier, prezzo").fetchall()]
    db.close()
    key=category_key(category)
    out=[]
    for r in rows:
        if category_key(r.get("categoria","")) != key:
            continue
        rv=r.get("volume_cl")
        if volume_cl is not None and rv is not None and abs(float(rv)-float(volume_cl)) > 0.01:
            continue
        out.append(r)
    return out


@app.get("/", response_class=HTMLResponse)
def home(request: Request):
    return templates.TemplateResponse("index.html", {"request": request})


@app.get("/api/products")
def api_products(q: str = "", category: str = ""):
    db = conn()
    sql = "SELECT * FROM bonaccini_products WHERE 1=1"
    params=[]
    if q:
        sql += " AND descrizione LIKE ?"; params.append(f"%{q}%")
    if category:
        sql += " AND categoria=?"; params.append(category)
    sql += " ORDER BY categoria, descrizione LIMIT 1000"
    out = [dict(r) for r in db.execute(sql, params).fetchall()]
    db.close(); return out


@app.get("/api/categories")
def api_categories():
    db = conn()
    cats = [r[0] for r in db.execute("""
        SELECT categoria FROM bonaccini_products
        UNION
        SELECT categoria FROM ortofrutticola_products
        ORDER BY categoria
    """).fetchall()]
    db.close(); return cats


@app.get("/api/stats")
def api_stats():
    db = conn()
    b = db.execute("SELECT COUNT(*) FROM bonaccini_products").fetchone()[0]
    o = db.execute("SELECT COUNT(*) FROM ortofrutticola_products").fetchone()[0]
    c = db.execute("SELECT COUNT(DISTINCT categoria) FROM bonaccini_products").fetchone()[0]
    suppliers={r["supplier"]:r["n"] for r in db.execute("SELECT supplier,COUNT(*) n FROM supplier_products GROUP BY supplier").fetchall()}
    db.close()
    return {"bonaccini":b,"ortofrutticola":o,"categorie":c,"suppliers":suppliers}


@app.get("/api/product/{pid}")
def api_product(pid: int):
    db=conn()
    b=db.execute("SELECT * FROM bonaccini_products WHERE id=?", (pid,)).fetchone()
    db.close()
    if not b: return JSONResponse({"error":"Prodotto non trovato"}, status_code=404)
    b=dict(b)
    return {"bonaccini":b,"matches":direct_matches_all(b),
            "alternatives": alternatives_all(b["categoria"], product_volume_cl(b)) if b["categoria"]!="Da classificare" else []}



def excel_sheets_and_headers(path):
    if load_workbook is None:
        raise RuntimeError("openpyxl non installato")
    wb=load_workbook(path, data_only=True, read_only=True)
    info=[]
    for ws in wb.worksheets:
        rows=[]
        for i,row in enumerate(ws.iter_rows(values_only=True)):
            vals=[clean_text(v) for v in row]
            if any(vals):
                rows.append(vals)
            if len(rows)>=8: break
        headers=rows[0] if rows else []
        info.append({"name":ws.title,"headers":headers,"preview":rows[:5]})
    wb.close()
    return info


def choose_column(headers, kind):
    hs=[clean_text(h).lower() for h in headers]
    if kind=='product':
        keys=('descrizione','prodotto','articolo','descr','voce','nome','merce')
    elif kind=='price':
        keys=('prezzo','price','costo','€','euro','listino')
    else:
        keys=('volume','formato','litri','litri','lt','cl','ml','quantita','quantità')
    for key in keys:
        for i,h in enumerate(hs):
            if key in h: return i
    return None


def import_generic_excel(path, supplier, sheet_name=None, product_col=None, price_col=None, volume_col=None):
    if load_workbook is None:
        raise RuntimeError("openpyxl non installato")
    supplier=clean_text(supplier)
    if not supplier: raise ValueError("Inserisci il nome del fornitore")
    wb=load_workbook(path, data_only=True, read_only=True)
    if sheet_name and sheet_name in wb.sheetnames:
        ws=wb[sheet_name]
    elif supplier.lower()=='conad' and 'CONAD VOLUME SEPARATO' in wb.sheetnames:
        ws=wb['CONAD VOLUME SEPARATO']
    else:
        ws=wb[wb.sheetnames[0]]
    rows=list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows: return 0,0,ws.title
    # trova la prima riga plausibile come intestazione
    header_idx=0
    for i,row in enumerate(rows[:15]):
        vals=[clean_text(v) for v in row]
        if any(('prezzo' in v.lower() or 'price' in v.lower() or 'prodotto' in v.lower() or 'descr' in v.lower() or 'voce' in v.lower()) for v in vals):
            header_idx=i; break
    headers=[clean_text(v) for v in rows[header_idx]]
    pc=product_col if product_col is not None else choose_column(headers,'product')
    pr=price_col if price_col is not None else choose_column(headers,'price')
    vc=volume_col if volume_col is not None else choose_column(headers,'volume')
    if pc is None or pr is None:
        raise ValueError('Non riesco a individuare automaticamente le colonne prodotto e prezzo. Usa l\'anteprima per scegliere le colonne.')
    db=conn(); cur=db.cursor(); inserted=0; updated=0
    for row in rows[header_idx+1:]:
        vals=list(row)
        if pc>=len(vals) or pr>=len(vals): continue
        desc=clean_text(vals[pc]); price=to_float(vals[pr])
        if not desc or price is None: continue
        vol_raw=vals[vc] if vc is not None and vc<len(vals) else None
        vol=volume_to_cl(vol_raw, desc)
        cat=infer_supplier_category(desc)
        existing=cur.execute('SELECT id FROM supplier_products WHERE supplier=? AND descrizione=? AND ((volume_cl IS NULL AND ? IS NULL) OR volume_cl=?)',(supplier,desc,vol,vol)).fetchone()
        unit=(f'{int(vol)} cl' if vol is not None and float(vol).is_integer() else (f'{vol:g} cl' if vol is not None else ''))
        if existing:
            cur.execute('UPDATE supplier_products SET prezzo=?, categoria=?, unita=?, source_file=? WHERE id=?',(price,cat,unit,path.name+' / '+ws.title,existing['id']))
            updated+=1
        else:
            cur.execute('INSERT INTO supplier_products (supplier,descrizione,categoria,prezzo,volume_cl,unita,source_file) VALUES (?,?,?,?,?,?,?)',(supplier,desc,cat,price,vol,unit,path.name+' / '+ws.title))
            inserted+=1
    db.commit(); db.close()
    return inserted,updated,ws.title

@app.post("/api/upload")
def upload(kind: str = Form(...), file: UploadFile = File(...)):
    target = UPLOADS / kind
    target.mkdir(parents=True, exist_ok=True)
    safe_name = Path(file.filename or "file").name
    name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_name}"
    path=target / name
    with path.open("wb") as f:
        shutil.copyfileobj(file.file, f)
    if kind == 'listini' and path.suffix.lower() in ('.xlsx','.xlsm'):
        try:
            return {"ok":True,"file":str(path.relative_to(BASE_DIR)),"sheets":excel_sheets_and_headers(path)}
        except Exception as e:
            return JSONResponse({"ok":False,"error":str(e)},status_code=400)
    return {"ok": True, "message": "File caricato.", "file": name}


@app.post("/api/import-supplier")
def import_supplier(
    supplier: str = Form(...),
    file: UploadFile = File(...),
    sheet: str = Form(""),
    product_col: str = Form(""),
    price_col: str = Form(""),
    volume_col: str = Form("")
):
    target=UPLOADS / 'listini'; target.mkdir(parents=True,exist_ok=True)
    safe_name=Path(file.filename or 'listino.xlsx').name
    name=f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{safe_name}"
    path=target/name
    with path.open('wb') as f: shutil.copyfileobj(file.file,f)
    try:
        def idx(x): return int(x) if str(x).strip()!='' else None
        ins,upd,used_sheet=import_generic_excel(path,supplier,sheet or None,idx(product_col),idx(price_col),idx(volume_col))
        return {"ok":True,"supplier":clean_text(supplier),"sheet":used_sheet,"inserted":ins,"updated":upd}
    except Exception as e:
        return JSONResponse({"ok":False,"error":str(e)},status_code=400)


@app.post("/api/category")
def update_category(product_id: int = Form(...), categoria: str = Form(...)):
    db = conn()
    db.execute("UPDATE bonaccini_products SET categoria=? WHERE id=?", (categoria, product_id))
    db.commit(); db.close()
    return RedirectResponse("/", status_code=303)


@app.post("/api/orders")
def create_order(lines: str = Form(...), note: str = Form("")):
    db = conn(); cur = db.cursor()
    cur.execute("INSERT INTO orders(created_at,note) VALUES(?,?)", (datetime.now().isoformat(timespec="seconds"), note))
    oid = cur.lastrowid
    for raw in lines.splitlines():
        if not raw.strip(): continue
        parts = raw.replace(",", ";").split(";")
        if len(parts) < 2: continue
        try:
            pid = int(parts[0].strip()); qty = float(parts[1].strip())
        except ValueError:
            continue
        b = db.execute("SELECT * FROM bonaccini_products WHERE id=?", (pid,)).fetchone()
        if not b: continue
        b=dict(b); matches=direct_matches_all(b)
        supplier="Bonaccini"; price=b["prezzo"]
        candidates=[(s,r["prezzo"]) for s,r in matches.items() if r.get("prezzo") is not None]
        if candidates and (price is None or min(p for _,p in candidates)<price):
            supplier,price=min(candidates,key=lambda x:x[1])
        cur.execute("INSERT INTO order_lines(order_id,descrizione,categoria,quantita,fornitore,prezzo) VALUES(?,?,?,?,?,?)", (oid,b["descrizione"],b["categoria"],qty,supplier,price))
    db.commit(); db.close()
    return RedirectResponse(f"/orders/{oid}", status_code=303)


@app.get("/orders/{oid}", response_class=HTMLResponse)
def order_page(request: Request, oid: int):
    db = conn()
    order = db.execute("SELECT * FROM orders WHERE id=?", (oid,)).fetchone()
    lines = db.execute("SELECT * FROM order_lines WHERE order_id=?", (oid,)).fetchall()
    db.close()
    return templates.TemplateResponse("order.html", {"request": request, "order": order, "lines": lines})
